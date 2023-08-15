from openpyxl import load_workbook
import os

# load input spreadsheet
files_in_directory = os.listdir("src/input")
for file in files_in_directory:
  file = "src/input/" + file
  wb = load_workbook(filename = file, read_only=True, data_only=True)
  break

# select worksheet
ws = wb['Project Info']
current_month = ws['D10'].value[0:3]
ws = wb['Shortname Summary']

# load output spreadsheet
files_in_directory = os.listdir("src/output")
for file in files_in_directory:
  file = "src/output/" + file
  wbx = load_workbook(filename = file)
  break
    
# select worksheet
wsx = wbx['Summary for Sponsors']

def main():
  months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
  actualColumns = getCurrentActuals(months, current_month)
  last_column = actualColumns[len(actualColumns)-1]
  forecastColumns = getCurrentForecasts(last_column)
  get_total = sumTotal()

  sumActuals(actualColumns)
  sumForecasts(forecastColumns)

  print("Program run...")

def getCurrentActuals(months, current_month):
  list = []
  counter = 0
  for col in ws.iter_rows(min_row=10, max_row=10, min_col=6, max_col=40):
    for cell in col:
      if cell.value == "Actuals":
        list.append(cell.column)
        counter += 1
      if months[counter-1] == current_month:
        return list
      
def getCurrentForecasts(last_column):
  list = []
  for col in ws.iter_rows(min_row=10, max_row=10, min_col=last_column+2, max_col=40):
    for cell in col:
      if cell.value == "Forecast":
        list.append(cell.column)
  return list


def sumActuals(actualColumns):
  sum = 0
  #run and manage
  for col in ws.iter_rows(min_row=96, max_row=104, min_col=6, max_col=40):
    for cell in col:
      if cell.row in [96, 97, 98, 102] and cell.column in actualColumns:
        sum += cell.value
  wsx["D3"] = int(sum/1000) 
  
  sum=0
  #Create Online Spanish-Language Application Experience
  for col in ws.iter_rows(min_row=96, max_row=104, min_col=6, max_col=40):
    for cell in col:
      if cell.row in [99] and cell.column in actualColumns:
        sum += cell.value
  wsx["D7"] = int(sum/1000) 

  sum=0
  #Automate Web Based Broker Application Submission
  for col in ws.iter_rows(min_row=96, max_row=104, min_col=6, max_col=40):
    for cell in col:
      if cell.row in [100] and cell.column in actualColumns:
        sum += cell.value
  wsx["D4"] = int(sum/1000) 

  sum=0
  #Paperless Delivery for KPIF Correspondences
  for col in ws.iter_rows(min_row=96, max_row=104, min_col=6, max_col=40):
    for cell in col:
      if cell.row in [101] and cell.column in actualColumns:
        sum += cell.value
  wsx["D5"] = int(sum/1000) 

  sum=0
  #Enhanced Direct Enrollment Single Sign On
  for col in ws.iter_rows(min_row=96, max_row=104, min_col=6, max_col=40):
    for cell in col:
      if cell.row in [103] and cell.column in actualColumns:
        sum += cell.value
    wsx["D6"] = int(sum)

  sum=0
 # Other CRs and Enhancements
  for col in ws.iter_rows(min_row=96, max_row=104, min_col=6, max_col=40):
    for cell in col:
      if cell.row in [104] and cell.column in actualColumns:
        sum += cell.value     
  wsx["D8"] = int(sum/1000) 

  wbx.save(filename=file)


def sumForecasts(forecastColumns):   
  sum = 0
  #run and manage
  for col in ws.iter_rows(min_row=96, max_row=104, min_col=6, max_col=40):
    for cell in col:
      if cell.row in [96, 97, 98, 102] and cell.column in forecastColumns:
        sum += cell.value
  wsx["E3"] = int(sum/1000) 
  
  sum=0
  #Create Online Spanish-Language Application Experience
  for col in ws.iter_rows(min_row=96, max_row=104, min_col=6, max_col=40):
    for cell in col:
      if cell.row in [99] and cell.column in forecastColumns:
        sum += cell.value
  wsx["E7"] = int(sum/1000) 

  sum=0
  #Automate Web Based Broker Application Submission
  for col in ws.iter_rows(min_row=96, max_row=104, min_col=6, max_col=40):
    for cell in col:
      if cell.row in [100] and cell.column in forecastColumns:
        sum += cell.value
  wsx["E4"] = int(sum/1000) 

  sum=0
  #Paperless Delivery for KPIF Correspondences
  for col in ws.iter_rows(min_row=96, max_row=104, min_col=6, max_col=40):
    for cell in col:
      if cell.row in [101] and cell.column in forecastColumns:
        sum += cell.value
  wsx["E5"] = int(sum/1000) 

  sum=0
  #Enhanced Direct Enrollment Single Sign On
  for col in ws.iter_rows(min_row=96, max_row=104, min_col=6, max_col=40):
    for cell in col:
      if cell.row in [103] and cell.column in forecastColumns:
        sum += cell.value
  wsx["E6"] = int(sum/1000) 

  sum=0
 # Other CRs and Enhancements
  for col in ws.iter_rows(min_row=96, max_row=104, min_col=6, max_col=40):
    for cell in col:
      if cell.row in [104] and cell.column in forecastColumns:
        sum += cell.value     
  wsx["E8"] = int(sum/1000) 

  wbx.save(filename=file)


def sumTotal():
  ws = wb['Project Info']
  
  wsx["C3"] = int((ws.cell(row=36, column=6).value + ws.cell(row=37, column=6).value + 
                  ws.cell(row=38, column=6).value + ws.cell(row=42, column=6).value)/1000)
  wsx["C4"] = int(ws.cell(row=40, column=6).value)/1000
  wsx["C5"] = int(ws.cell(row=41, column=6).value)/1000
  wsx["C6"] = int(ws.cell(row=43, column=6).value)/1000
  wsx["C7"] = int(ws.cell(row=39, column=6).value)/1000
  wsx["C8"] = int(ws.cell(row=44, column=6).value)/1000

  wbx.save(filename=file)
  



main()